from openpyxl import Workbook, load_workbook
import requests
from bs4 import BeautifulSoup
import sqlite3
import time
import hashlib
import os
import re

# =========================
# CONFIG (Railway)
# =========================
#TOKEN = "8643526470:AAEqdTHnTz1YNRP7aSntyS93ULbTnChkMjA"
#CHAT_ID = "1169857949"

TOKEN = "8643526470:AAEqdTHnTz1YNRP7aSntyS93ULbTnChkMjA"
CHAT_ID = "1169857949"

# =========================
# BANCO
# =========================
conn = sqlite3.connect("corridas.db")
cursor = conn.cursor()

cursor.execute("""
CREATE TABLE IF NOT EXISTS eventos (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    hash TEXT UNIQUE,
    titulo TEXT,
    link TEXT,
    score INTEGER
)
""")

# =========================
# REGRAS
# =========================
PALAVRAS_DF = ["brasília", "df", "taguatinga", "ceilândia"]
PALAVRAS_GRATUITO = ["gratuito", "gratuita", "free", "sem custo"]
PALAVRAS_ORGAOS = ["gdf", "sesc", "sesi"]

# =========================
# FUNÇÕES
# =========================

def calcular_score(texto):
    t = texto.lower()
    score = 0

    if any(p in t for p in PALAVRAS_DF):
        score += 30

    if any(p in t for p in PALAVRAS_GRATUITO):
        score += 100

    if any(p in t for p in PALAVRAS_ORGAOS):
        score += 40

    if "r$" in t:
        valores = re.findall(r"\d+", t)
        if valores:
            preco = int(valores[0])
            if preco <= 100:
                score += 80
            else:
                score -= 50

    return score


def eh_df(texto):
    t = texto.lower()
    return any(p in t for p in PALAVRAS_DF)


def gerar_hash(titulo, link):
    return hashlib.md5((titulo + link).encode()).hexdigest()


def enviar_telegram(msg):
    url = f"https://api.telegram.org/bot{TOKEN}/sendMessage"
    response = requests.post(url, data={
        "chat_id": CHAT_ID,
        "text": msg,
        "parse_mode": "Markdown"
    })
    print("Resposta Telegram:", response.text)


# =========================
# SCRAPER
# =========================

def buscar_eventos():
    url = "https://www.corridasbr.com.br/df/calendario.asp"
    eventos = []

    try:
        response = requests.get(url, timeout=10)
        soup = BeautifulSoup(response.text, "html.parser")

        for item in soup.find_all("a"):
            titulo = item.text.strip()
            link = item.get("href")

            if titulo and link:
                if not link.startswith("http"):
                    link = "https://www.corridasbr.com.br/df/" + link

                eventos.append((titulo, link))

    except Exception as e:
        print("Erro ao buscar eventos:", e)

    return eventos


# =========================
# EXTRAÇÃO AVANÇADA
# =========================

def extrair_detalhes(link):
    try:
        response = requests.get(link, timeout=10)
        soup = BeautifulSoup(response.text, "html.parser")

        texto = soup.get_text().lower()

        # 📅 DATA
        data = "N/A"
        datas = re.findall(r"\d{1,2}/\d{1,2}/\d{2,4}", texto)
        if datas:
            data = datas[0]

        # 💰 PREÇO
        preco = "N/A"
        valores = re.findall(r"r\$\s*\d+", texto)
        if valores:
            preco = valores[0]

        # 📏 DISTÂNCIA
        distancia = "N/A"
        if "5k" in texto:
            distancia = "5km"
        elif "10k" in texto:
            distancia = "10km"
        elif "21k" in texto:
            distancia = "21km"

        # 🎽 KIT
        kit = []
        if "camiseta" in texto:
            kit.append("Camiseta")
        if "medalha" in texto:
            kit.append("Medalha")
        if "boné" in texto:
            kit.append("Boné")

        kit = ", ".join(kit) if kit else "N/A"

        return data, preco, distancia, kit

    except Exception as e:
        print("Erro ao extrair detalhes:", e)
        return "N/A", "N/A", "N/A", "N/A"


# =========================
# EXCEL
# =========================

def salvar_excel(titulo, link, score, data, preco, distancia, kit):

    arquivo = "corridas.xlsx"

    if not os.path.exists(arquivo):
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Título", "Link", "Data", "Preço",
            "Distância", "Kit", "Score"
        ])
        wb.save(arquivo)

    wb = load_workbook(arquivo)
    ws = wb.active

    ws.append([
        titulo,
        link,
        data,
        preco,
        distancia,
        kit,
        score
    ])

    wb.save(arquivo)


# =========================
# PROCESSAMENTO
# =========================

def processar():
    eventos = buscar_eventos()

    print(f"🔍 Encontrados: {len(eventos)}")

    for titulo, link in eventos:

        if not eh_df(titulo):
            continue

        score = calcular_score(titulo)

        if score < 50:
            continue

        h = gerar_hash(titulo, link)

        try:
            cursor.execute(
                "INSERT INTO eventos (hash, titulo, link, score) VALUES (?, ?, ?, ?)",
                (h, titulo, link, score)
            )
            conn.commit()

            print("🔎 Analisando:", titulo)

            data, preco, distancia, kit = extrair_detalhes(link)

            print("📊 SALVANDO NO EXCEL:", titulo)

            salvar_excel(titulo, link, score, data, preco, distancia, kit)

            mensagem = f"""
🚨 *CORRIDA ENCONTRADA*

🏷 {titulo}
📅 Data: {data}
💰 Preço: {preco}
📏 Distância: {distancia}
🎽 Kit: {kit}

🔗 {link}
"""

            enviar_telegram(mensagem)

        except sqlite3.IntegrityError:
            print("⚠️ Evento já registrado, ignorando...") 

            
# =========================
# LOOP
# =========================

while True:
    print("🔄 Rodando...")
    processar()
    time.sleep(600)